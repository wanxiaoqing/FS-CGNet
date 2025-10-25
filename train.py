import datetime
import os
import random
import re
import sys

import numpy as np
import pandas as pd
import scipy.io as sio
from openpyxl import load_workbook
from sklearn.decomposition import PCA
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report, cohen_kappa_score
import torch
import torch.nn as nn
import torch.optim as optim
from operator import truediv
from utils import get_cls_map
import time
from torch.optim.lr_scheduler import CosineAnnealingLR

from main import FSCGNet

CLASS_IP = ['Alfalfa', 'Corn-notill', 'Corn-mintill', 'Corn', 'Grass-pasture', 'Grass-trees',
            'Grass-pasture-mowed', 'Hay-windrowed', 'Oats', 'Soybean-notill', 'Soybean-mintill',
            'Soybean-clean', 'Wheat', 'Woods', 'Buildings-Grass-Trees-Drives', 'Stone-Steel-Towers']
CLASS_BA = ['Alfalfa', 'Corn-notill', 'Corn-mintill', 'Corn', 'Grass-pasture', 'Grass-trees',
            'Grass-pasture-mowed', 'Hay-windrowed', 'Oats', 'Soybean-notill', 'Soybean-mintill',
            'Soybean-clean', 'Wheat', 'Woods']
CLASS_PU = ['Asphalt', 'Meadows', 'Gravel', 'Trees', 'Painted metal sheets', 'Bare Soil',
            'Bitumen', 'Self-Blocking Bricks', 'Shadows']
CLASS_HC = ["Strawberry", "Grass", "Cowpea", "Red roof", "Soybean", "Gray roof", "Sorghum", "Plastic",
            "Water spinach", "Bare soil", "Watermelon", "Road", "Greens", "Bright", "Trees", "Water", ]
CLASS_HH = ["Red roof", "Brassica chinensis", "Road", "Small Brassica chinensis", "Bare soil", "Lactuca sativa",
            "Cotton", "Celtuce", "Cotton firewood", "Film covered lettuce", "Rape", "Romaine lettuce",
            "Chinese cabbage", "Carrot", "Pakchoi", "White radish", "Cabbage", "Garlic sprout", "Tuber mustard",
            "Broad bean", "Brassica parachinensis", "Tree", ]
CLASS_HU = ["Red roof", "Brassica chinensis", "Road", "Small Brassica chinensis", "Bare soil", "Lactuca sativa",
            "Cotton", "Celtuce", "Cotton firewood", "Film covered lettuce", "Rape", "Romaine lettuce",
            "Chinese cabbage", "Carrot", "Pakchoi", ]
CLASS_LK = ["Corn", "Cotton", "Sesame", "Broad-leaf soybean", "Narrow-leaf soybean", "Rice", "Water",
            "Roads and houses", "Mixed weed", ]
CLASS_SA = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16"]
CLASS_KSC = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13"]
CLASS_XA = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
            "20"]
CLASS_PC = ["1", "2", "3", "4", "5", "6", "7", "8", "9"]
name_list = {
    'IP': ["Indian_pines_corrected.mat", "Indian_pines_gt.mat", "indian_pines_corrected", "indian_pines_gt", CLASS_IP],
    'PU': ["PaviaU.mat", "PaviaU_gt.mat", "paviaU", "paviaU_gt", CLASS_PU],
    'KSC': ["KSC.mat", "KSC_gt.mat", "KSC", "KSC_gt", CLASS_KSC],
    'BA': ["Botswana.mat", "Botswana_gt.mat", "Botswana", "Botswana_gt", CLASS_BA],
    'HU': ["Houston.mat", "Houston_gt.mat", "Houston", "Houston_gt", CLASS_HU],
    'HH': ["WHU_Hi_HongHu.mat", "WHU_Hi_HongHu_gt.mat", "WHU_Hi_HongHu", "WHU_Hi_HongHu_gt", CLASS_HH],
    'HC': ["WHU_Hi_HanChuan.mat", "WHU_Hi_HanChuan_gt.mat", "WHU_Hi_HanChuan", "WHU_Hi_HanChuan_gt", CLASS_HC],
    'LK': ["WHU_Hi_LongKou.mat", "WHU_Hi_LongKou_gt.mat", "WHU_Hi_LongKou", "WHU_Hi_LongKou_gt", CLASS_LK],
    'SA': ["Salinas_corrected.mat", "Salinas_gt.mat", "salinas_corrected", "salinas_gt", CLASS_SA],
    'XA': ["XiongAn.mat", "XiongAn_gt.mat", "XiongAn", "XiongAn_gt", CLASS_XA],
    'PC': ["Pavia.mat", "Pavia_gt.mat", "pavia", "pavia_gt", CLASS_PC],
}

conv_params = {
    9: (3, 2, 1),
    11: (3, 2, 0),
    13: (3, 3, 1),
    15: (5, 3, 1),
    17: (5, 3, 0),
    19: (5, 4, 1),
    21: (5, 4, 0),
    23: (5, 5, 1),
    25: (5, 5, 0)
}

T_max = 150


def get_numpy_tensor_memory(tensor: np.ndarray) -> float:
    """
    计算 NumPy 张量的内存占用（以 GB 为单位）

    Args:
    tensor (np.ndarray): 输入的 NumPy 张量

    Returns:
    float: 张量占用的内存大小（GB）
    """
    # 获取张量的总字节数
    size_in_bytes = tensor.nbytes

    # 转换为 GB
    size_in_gb = size_in_bytes / (1024 ** 3)
    return size_in_gb


def get_torch_tensor_memory(tensor: torch.Tensor) -> float:
    """
    计算 PyTorch 张量的内存占用（以 GB 为单位）

    Args:
    tensor (torch.Tensor): 输入的 PyTorch 张量

    Returns:
    float: 张量占用的内存大小（GB）
    """
    # 获取每个元素的字节大小
    size_in_bytes = tensor.element_size() * tensor.numel()

    # 转换为 GB
    size_in_gb = size_in_bytes / (1024 ** 3)
    return size_in_gb


# 修改get_torch_tensor_memory以计算Dataset对象中的数据内存占用
def get_dataset_memory(dataset) -> float:
    # 估算数据集内存占用
    # 获取数据集的第一个batch并计算其内存
    first_batch = next(iter(dataset))
    data, labels = first_batch  # 获取数据和标签

    # 计算数据部分的内存
    data_memory = get_torch_tensor_memory(data)
    labels_memory = get_torch_tensor_memory(labels)

    # 返回数据和标签的总内存
    return data_memory + labels_memory


def save_result_to_excel(oas, aas, kappas, parameters, folder_name, module_name, dataset_name):
    """
    保存超参数和运行结果到 Excel 文件，并在不同批次间插入空行。

    Args:
        oas (list): OA 列表。
        aas (list): AA 列表。
        kappas (list): Kappa 列表。
        parameters (dict): 超参数，键为参数名称，值为参数值。
        folder_name (str): 文件夹名称。
        module_name (str): 模块名称。
        dataset_name (str): 数据集名称。
    """

    # 定义文件夹路径和文件名
    folder_path = f"cls_result/{folder_name}/{module_name}/{dataset_name}/"
    file_name = f"{folder_name}_{module_name}_{dataset_name}_训练精度.xlsx"
    file_path = os.path.join(folder_path, file_name)
    sheet_name = "Sheet1"  # 工作表名称，固定为 Sheet1，可根据需要调整

    # 确保文件夹路径存在
    os.makedirs(folder_path, exist_ok=True)

    # 创建结果数据
    results = {
        "Index": list(range(1, len(oas) + 1)) + ["Average"],
        "OA (%)": oas + [np.mean(oas)],
        "AA (%)": aas + [np.mean(aas)],
        "Kappa (%)": kappas + [np.mean(kappas)]
    }

    # 将超参数和结果整合到一个 DataFrame
    param_df = pd.DataFrame([parameters])
    results_df = pd.DataFrame(results)
    combined_df = pd.concat([param_df, results_df], axis=1)

    # 添加分隔空行
    empty_row = pd.DataFrame([[""] * combined_df.shape[1]], columns=combined_df.columns)
    combined_df = pd.concat([combined_df, empty_row], ignore_index=True)

    # 检查文件是否存在
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            workbook = load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                # 加载现有数据并插入分隔空行
                existing_data = pd.read_excel(file_path, sheet_name=sheet_name)
                combined_df = pd.concat([existing_data, empty_row, combined_df], ignore_index=True)
            combined_df.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        # 文件不存在时，创建新文件
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            combined_df.to_excel(writer, index=False, sheet_name=sheet_name)

    print(f"数据已成功保存到 {file_path} 的工作表 {sheet_name}")


def loadData(name):
    data_path = os.path.join(r"data", name_list[name][0])
    gt_path = os.path.join(r"data", name_list[name][1])

    data = sio.loadmat(data_path)[name_list[name][2]]
    label = sio.loadmat(gt_path)[name_list[name][3]]

    return data, label


# 对高光谱数据 X 应用 PCA 变换
def applyPCA(X, numComponents):
    newX = np.reshape(X, (-1, X.shape[2]))
    pca = PCA(n_components=numComponents, whiten=True)
    newX = pca.fit_transform(newX)
    newX = np.reshape(newX, (X.shape[0], X.shape[1], numComponents))

    return newX


# 对单个像素周围提取 patch 时，边缘像素就无法取了，因此，给这部分像素进行 padding 操作
def padWithZeros(X, margin=2):
    # 创建一个新的零张量，尺寸比原始张量大
    # 这里假设输入 X 是一个三维张量，shape 为 [C, H, W]
    newX = torch.zeros((X.shape[0], X.shape[1] + 2 * margin, X.shape[2] + 2 * margin), device=X.device)

    # 设置偏移量
    x_offset = margin
    y_offset = margin

    # 将原始数据填充到新张量中
    newX[:, x_offset:X.shape[1] + x_offset, y_offset:X.shape[2] + y_offset] = X

    return newX


def extract_single_pixel_patch(x_coor, y_coor, X, margin):
    x_min = x_coor
    y_min = y_coor
    x_max = x_coor + 2 * margin + 1
    y_max = y_coor + 2 * margin + 1

    # 提取 patch
    patch = X[:, x_min:x_max, y_min:y_max]
    # for c in range(patch.size(0)):
    #     for i in range(patch.size(1)):  # 遍历高度
    #         for j in range(patch.size(2)):  # 遍历宽度
    #             print(f"tensor[{c}, {i}, {j}] = {patch[c, i, j].item()}")

    patch = patch.unsqueeze(0)

    return patch


def extract_patches(X, train_coords, patch_size=11):
    # 确定补丁的一半大小
    half_size = patch_size // 2
    train_num = train_coords.shape[0]
    num_bands = X.shape[0]

    # 初始化结果张量
    x_train = torch.zeros((train_num, num_bands, patch_size, patch_size), dtype=X.dtype)

    # 遍历每个坐标点
    for i, (row, col) in enumerate(train_coords):
        # 提取补丁的起始和结束索引
        row_start = row
        row_end = row + 2 * half_size + 1
        col_start = col
        col_end = col + 2 * half_size + 1

        # 提取补丁
        patch = X[:, row_start:row_end, col_start:col_end]

        # 确保补丁大小为 11x11，忽略不满足条件的样本
        if patch.shape[1:] == (patch_size, patch_size):
            x_train[i] = patch
        else:
            raise ValueError(f"Patch at index {i} has incorrect shape: {patch.shape[1:]}. Check boundary conditions.")

    return x_train


# 在每个像素周围提取 patch ，然后创建成符合 keras 处理的格式
def createImageCubes(X, y, windowSize=5, removeZeroLabels=False):
    labels_coordinates = np.array(list(zip(*np.nonzero(y))))
    labels = y[labels_coordinates[:, 0], labels_coordinates[:, 1]]
    labels = labels - 1

    return labels_coordinates, labels


# def splitTrainTestSet(X, y, testRatio, randomState=random.randint(0, 10000000)):
def splitTrainTestSet(X, y, testRatio, randomState=345):
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=testRatio, random_state=randomState, stratify=y)
    return X_train, X_test, y_train, y_test


def create_data_loader():
    # 读入数据
    X, y = loadData(DATA_SET_NAME)

    print('Hyperspectral data shape: ', X.shape)
    print('Label shape: ', y.shape)

    # 打印 NumPy 数据内存占用
    # print(f"Memory usage of X (NumPy array): {get_numpy_tensor_memory(X):.6f} GB")
    # print(f"Memory usage of y (NumPy array): {get_numpy_tensor_memory(y):.6f} GB")

    print('\n... ... PCA tranformation ... ...')
    X = applyPCA(X, numComponents=pca_components)
    print('Data shape after PCA: ', X.shape)

    # 打印 PCA 后的数据内存占用
    # print(f"Memory usage of X_pca (NumPy array): {get_numpy_tensor_memory(X):.6f} GB")

    print('\n... ... create data cubes ... ...')
    # print("00", X.shape)
    X = X.transpose(2, 0, 1)
    # print("11", X.shape)
    x_coor, y_all = createImageCubes(X, y, windowSize=PATCH_SIZE)

    X = torch.FloatTensor(X).to("cuda:0")
    y = torch.LongTensor(y).to("cuda:0")

    print('Data cube X shape: ', X.shape)
    print('Data cube y shape: ', y.shape)

    # 打印数据立方体后的内存占用
    # print(f"Memory usage of X_pca (NumPy array after creating cubes): {get_numpy_tensor_memory(x_coor):.6f} GB")
    # print(f"Memory usage of y_all (NumPy array after creating cubes): {get_numpy_tensor_memory(y_all):.6f} GB")

    print('\n... ... create train & test data ... ...')
    Xtrain_coor, Xtest_coor, ytrain_coor, ytest_coor = splitTrainTestSet(x_coor, y_all, test_ratio)
    # print("444", X.shape)
    X = padWithZeros(X, MARGIN)
    X_train = extract_patches(X, Xtrain_coor, PATCH_SIZE)
    # print("555", X_train.shape)
    # 第 1 个维度（从 0 开始计数）插入一个大小为 1 的新维度
    X_train = X_train.unsqueeze(1)  # HybridSN 使用时要注释掉  CVSSN 使用时要注释掉
    # print("666", X_train.shape)
    # print('Xtrain shape: ', Xtrain_coor.shape)
    # print('Xtest  shape: ', Xtest_coor.shape)

    # 创建train_loader和 test_loader
    all___set = TestDS(x_coor, y_all, X, y)
    train_set = TrainDS(X_train, ytrain_coor, )
    test__set = TestDS(Xtest_coor, ytest_coor, X, y)

    all_data_loader = torch.utils.data.DataLoader(dataset=all___set, batch_size=BATCH_SIZE_TRAIN, shuffle=False, )
    train_loader = torch.utils.data.DataLoader(dataset=train_set, batch_size=BATCH_SIZE_TRAIN, shuffle=True, )
    # num_workers=6,  # 根据 CPU 核心数设置
    # pin_memory=True,  # 加速数据传输到 GPU
    # prefetch_factor=1, )  # 每个 worker 预加载的批次数量
    test_loader = torch.utils.data.DataLoader(dataset=test__set, batch_size=BATCH_SIZE_TRAIN, shuffle=False, )

    print('\n... ... Creating training and testing datasets is complete ... ...')

    return train_loader, test_loader, all_data_loader, y


# Training dataset
class TrainDS(torch.utils.data.Dataset):
    def __init__(self, X_train, y_label, ):
        self.len = X_train.shape[0]
        self.X_train = X_train
        self.y_label = y_label

    def __getitem__(self, index):
        # 根据索引返回数据和对应的标签
        return self.X_train[index], self.y_label[index]

    def __len__(self):
        # 返回文件数据的数目
        return self.len


# Testing dataset
class TestDS(torch.utils.data.Dataset):
    def __init__(self, label_coor, y_label, X, y):
        self.len = label_coor.shape[0]

        self.label_coor = label_coor
        self.y_label = y_label

        self.X = X
        self.y = y

    def __getitem__(self, index):
        x_coor = self.label_coor[index][0]
        y_coor = self.label_coor[index][1]

        patch = extract_single_pixel_patch(x_coor, y_coor, self.X, MARGIN)
        # print("0", patch.shape)

        # 调整维度顺序，从 [30, 1, 9, 9] 到 [1, 30, 9, 9]，用于MASSF模型
        # patch = patch.permute(1, 0, 2, 3)
        # print("1", patch.shape)  # 打印调整后的形状 [1, 30, 9, 9]

        # 根据索引返回数据和对应的标签
        return patch, self.y_label[index]

    def __len__(self):
        # 返回文件数据的数目
        return self.len


def count_parameters(model):
    return sum(p.numel() for p in model.parameters() if p.requires_grad)


def train(train_loader, test_loader, epochs):
    # 使用GPU训练，可以在菜单 "代码执行工具" -> "更改运行时类型" 里进行设置
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

    net = FSCGNet(in_channels=1, num_classes=NUM_classes).to(device)

    # print(f"Number of parameters: {count_parameters(net):,}")

    # 交叉熵损失函数
    criterion = nn.CrossEntropyLoss()
    # 初始化优化器
    optimizer = optim.Adam(net.parameters(), lr=learning_rate)
    scheduler = CosineAnnealingLR(optimizer, T_max=T_max)
    # 开始训练
    total_loss = 0
    for epoch in range(epochs):
        net.train()
        for i, (data, target) in enumerate(train_loader):
            data, target = data.to(device), target.to(device)
            """     
            以下四行用于CVSSN的调用，不使用CVSSN时要注释掉       
            x_spa = data.permute(0, 2, 3, 1)
            center = Patch_Size // 2
            x_spe = data[:, :, center, center]  
            outputs = net(x_spa, x_spe)"""
            # x_spa = data.permute(0, 2, 3, 1)
            # center = Patch_Size // 2
            # x_spe = data[:, :, center, center]
            # outputs = net(x_spa, x_spe)
            # 正向传播 +　反向传播 + 优化
            # 通过输入得到预测的输出
            outputs = net(data)  # 单个参数的模型使用，传入一个参数
            """
            outputs = net(data, data) 用于M2Fnet和MFT模型，不用则注释掉
            """
            # outputs = net(data, data)

            # 计算损失函数
            loss = criterion(outputs, target)
            # 优化器梯度归零
            optimizer.zero_grad()
            # 反向传播
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
        scheduler.step()
        # print('\r[Epoch: %d]  [loss avg: %.4f]  [current loss: %.6f]' % (
        #     epoch + 1, total_loss / (epoch + 1), loss.item(),), end="")
        print('[Epoch: %d]   [loss avg: %.4f]   [current loss: %.4f]' % (epoch + 1,
                                                                         total_loss / (epoch + 1),
                                                                         loss.item()))

        # y_pred_test, y_test = test(device, net, test_loader)
        # classification, oa, confusion, each_acc, aa, kappa = acc_reports(y_test, y_pred_test)
        # print(f"OA: {oa:.2f}%, AA: {aa:.2f}%, Kappa: {kappa:.2f}%,")

    # print('Finished Training', end="")

    return net, device


def test(device, net, test_loader):
    count = 0
    # 模型测试
    net.eval()
    y_pred_test = 0
    y_test = 0
    for inputs, labels in test_loader:
        inputs = inputs.to(device)
        outputs = net(inputs)  # 单个参数的模型使用，传入一个参数
        """
        以下四行用于CVSSN的调用，不使用CVSSN时要注释掉 
        x_spa = inputs.permute(0, 2, 3, 1)
        center = Patch_Size // 2
        x_spe = inputs[:, :, center, center]
        outputs = net(x_spa, x_spe)"""
        # x_spa = inputs.permute(0, 2, 3, 1)
        # center = Patch_Size // 2
        # x_spe = inputs[:, :, center, center]
        # outputs = net(x_spa, x_spe)
        """
        outputs = net(inputs, inputs) 用于M2Fnet,MFT，不用需要注释掉
        """
        # outputs = net(inputs, inputs)
        outputs = np.argmax(outputs.detach().cpu().numpy(), axis=1)
        if count == 0:
            y_pred_test = outputs
            y_test = labels
            count = 1
        else:
            y_pred_test = np.concatenate((y_pred_test, outputs))
            y_test = np.concatenate((y_test, labels))

    print('Finished Testing')
    return y_pred_test, y_test


def AA_andEachClassAccuracy(confusion_matrix):
    list_diag = np.diag(confusion_matrix)
    list_raw_sum = np.sum(confusion_matrix, axis=1)
    each_acc = np.nan_to_num(truediv(list_diag, list_raw_sum))
    average_acc = np.mean(each_acc)
    return each_acc, average_acc


def acc_reports(y_test, y_pred_test):
    classification = classification_report(y_test, y_pred_test, digits=4, target_names=name_list[DATA_SET_NAME][4])
    oa = accuracy_score(y_test, y_pred_test)
    confusion = confusion_matrix(y_test, y_pred_test)
    each_acc, aa = Precision_andEachClassAccuracy(confusion)  # 使用 Precision 替代 Recall 计算 AA
    kappa = cohen_kappa_score(y_test, y_pred_test)

    return classification, oa * 100, confusion, each_acc * 100, aa * 100, kappa * 100


def Precision_andEachClassAccuracy(confusion_matrix):
    list_diag = np.diag(confusion_matrix)  # True Positives (TP)
    list_col_sum = np.sum(confusion_matrix, axis=0)  # Sum of predictions per class (TP + FP)
    each_precision = np.nan_to_num(list_diag / list_col_sum)  # Precision for each class
    average_precision = np.mean(each_precision)  # Average Precision (AA)
    return each_precision, average_precision


def main():
    train_loader, test_loader, all_data_loader, y_all = create_data_loader()
    tic1 = time.perf_counter()
    net, device = train(train_loader, test_loader, epochs=Epochs)
    # 只保存模型参数
    # torch.save(net.state_dict(), f'./cls_params/LS_net_{DATA_SET_NAME}_{PATCH_SIZE}_params.pth')
    toc1 = time.perf_counter()
    tic2 = time.perf_counter()
    y_pred_test, y_test = test(device, net, test_loader)
    toc2 = time.perf_counter()
    # 评价指标
    classification, oa, confusion, each_acc, aa, kappa = acc_reports(y_test, y_pred_test)
    classification = str(classification)
    Training_Time = toc1 - tic1
    Test_time = toc2 - tic2

    # aa = float(re.search(r'macro avg\s+([\d\.]+)', classification).group(1)) * 100

    print(
        f"OA: {oa:.2f}%, AA: {aa:.2f}%, Kappa: {kappa:.2f}%, Training Time: {toc1 - tic1:.2f}s, Test Time: {toc2 - tic2:.2f}s")

    file_name = f"cls_result/{Folder_Name}/{Module_Name}/{DATA_SET_NAME}/{Patch_Size}/LR{int(round(learning_rate * 10000))}/TR_{int(Test_Ratio * 100)}_{Index}.txt"
    os.makedirs(os.path.dirname(file_name), exist_ok=True)
    with open(file_name, 'w') as x_file:
        x_file.write('\nTimestamp: {}\n'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        x_file.write('Dataset name: {}\n'.format(name_list[DATA_SET_NAME][2]))
        x_file.write('Batch size: {}\n'.format(BATCH_SIZE_TRAIN))
        x_file.write('Patch size: {}\n'.format(PATCH_SIZE))
        x_file.write('Learning rate: {}\n'.format(learning_rate))
        x_file.write('T_max: {}\n'.format(T_max))
        x_file.write('PCA dimensions: {}\n'.format(pca_components))
        x_file.write('{} Training_Time (s)'.format(Training_Time))
        x_file.write('\n')
        x_file.write('{} Test_time (s)'.format(Test_time))
        x_file.write('\n')
        x_file.write('{} Kappa accuracy (%)'.format(kappa))
        x_file.write('\n')
        x_file.write('{} Overall accuracy (%)'.format(oa))
        x_file.write('\n')
        x_file.write('{} Average accuracy (%)'.format(aa))
        x_file.write('\n')
        x_file.write('{} \nEach accuracy (%)'.format(each_acc))
        x_file.write('\n')
        x_file.write('{}'.format(classification))
        x_file.write('\n')
        x_file.write('{}'.format(confusion))
        print('The training and testing data have been saved.')

    get_cls_map.get_cls_map(net, device, all_data_loader, y_all, Folder_Name, DATA_SET_NAME, Patch_Size, LearningRate,
                            Test_Ratio,
                            Index, Module_Name)
    return oa, aa, kappa


if __name__ == "__main__":
    Folder_Name = "对比3-18"

    Module_Name = "SSFTT"

    dataset_config = {
        # "BA": {
        #     "PATCH_SIZE": [17],
        #     "learning_rate": [0.0003],
        #     "test_ratio": [0.90],
        #     "index": [1],
        #     "NUM_classes": 14,
        #     "BATCH_SIZE_TRAIN": 32,
        #     "Epochs": 100,
        #     "pca_components": 30,
        # },
        # "PU": {
        #     "PATCH_SIZE": [15],
        #     "learning_rate": [0.0002],
        #     "test_ratio": [0.90],
        #     # "index": [1],
        #     # "test_ratio": [0.92, 0.94],
        #     "index": [1],
        #     "NUM_classes": 9,
        #     "BATCH_SIZE_TRAIN": 512,
        #     "Epochs": 100,
        #     "pca_components": 30,
        # },
        # "HC": {
        #     "PATCH_SIZE": [11],
        #     "learning_rate": [0.00001],
        #     "test_ratio": [0.90],
        #     "index": [5],
        #     # "test_ratio": [0.90, 0.92, 0.94, 0.96, 0.98],
        #     # "index": [1, 2, 3, 4, 5],
        #     "NUM_classes": 16,
        #     "BATCH_SIZE_TRAIN": 32,
        #     "Epochs": 100,
        #     # "pca_components": 274,
        #     "pca_components": 30
        # },
        # "HH": {
        #     "PATCH_SIZE": [11],
        #     "learning_rate": [0.00001],
        #     "test_ratio": [0.90],
        #     "index": [1, 2, 3, 4, 5],
        #     # "test_ratio": [0.90, 0.92, 0.94, 0.96, 0.98],
        #     # "index": [1, 2, 3, 4, 5],
        #     "NUM_classes": 22,
        #     "BATCH_SIZE_TRAIN": 32,
        #     "Epochs": 100,
        #     # "pca_components": 270,
        #     "pca_components": 30,
        # },
        "SA": {
            "PATCH_SIZE": [17],
            "learning_rate": [0.0002],
            # "test_ratio": [0.90],
            # "index": [1, 2, 3, 4, 5],
            "test_ratio": [0.90],
            "index": [1],
            "NUM_classes": 16,
            "BATCH_SIZE_TRAIN": 512,
            "Epochs": 100,
            "pca_components": 30
        },
        # "HU": {
        #     "PATCH_SIZE": [13],
        #     "learning_rate": [0.0004],
        #     # "test_ratio": [0.90],
        #     # "index": [1, 2, 3, 4, 5],
        #     "test_ratio": [0.90],
        #     "index": [1],
        #     "NUM_classes": 15,
        #     "BATCH_SIZE_TRAIN": 64,
        #     "Epochs": 100,
        #     "pca_components": 30
        # },
        # "IP": {
        #     "PATCH_SIZE": [15],
        #     "learning_rate": [0.0001],
        #     "test_ratio": [0.90],
        #     "index": [1],
        #     "NUM_classes": 16,
        #     "BATCH_SIZE_TRAIN": 64,
        #     "Epochs": 100,
        #     "pca_components": 30
        # },
        # "KSC": {
        #     "PATCH_SIZE": [17],
        #     "learning_rate": [0.0005],
        #     "test_ratio": [0.90],
        #     "index": [1],
        #     "NUM_classes": 13,
        #     "BATCH_SIZE_TRAIN": 64,
        #     "Epochs": 100,
        #     # "pca_components": 176,
        #     "pca_components": 30
        # },    ·
        # "XA": {
        #     "PATCH_SIZE": [5],
        #     "learning_rate": [0.0001],
        #     # "test_ratio": [0.98],
        #     "test_ratio": [0.90],
        #     "index": [1],
        #     "NUM_classes": 20,
        #     "BATCH_SIZE_TRAIN": 1024,
        #     "Epochs": 50,
        #     # "pca_components": 250,
        #     "pca_components": 30
        # },
        # "LK": {
        #     "PATCH_SIZE": [17],
        #     "learning_rate": [0.0003],
        #     "test_ratio": [0.90],
        #     "index": [1],
        #     "NUM_classes": 9,
        #     "BATCH_SIZE_TRAIN": 64,  # 512
        #     "Epochs": 100,
        #     "pca_components": 30
        # },
        # "PC": {
        #     "PATCH_SIZE": [17],
        #     "learning_rate": [0.0001],
        #     "test_ratio": [0.90],
        #     # "index": [1],
        #     # "test_ratio": [0.92, 0.94],
        #     "index": [1],
        #     "NUM_classes": 9,
        #     "BATCH_SIZE_TRAIN": 32,
        #     "Epochs": 100,
        #     "pca_components": 30,
        # },
    }

    # 修改后的循环
    for name, config in dataset_config.items():
        DATA_SET_NAME = name
        NUM_classes = config["NUM_classes"]  # 获取当前数据集的 NUM_classes
        BATCH_SIZE_TRAIN = config["BATCH_SIZE_TRAIN"]  # 从 config 中获取 BATCH_SIZE_TRAIN
        Epochs = config["Epochs"]  # 从 config 中获取 Epochs
        pca_components = config["pca_components"]

        for PATCH_SIZE in config["PATCH_SIZE"]:
            Patch_Size = PATCH_SIZE
            for learning_rate in config["learning_rate"]:
                LearningRate = learning_rate
                # for test_ratio in [0.90, 0.92, 0.94, 0.96, 0.98]:  # 0.90-0.98, 间隔0.02
                for test_ratio in config["test_ratio"]:  # 0.90-0.98, 间隔0.02
                    Test_Ratio = test_ratio
                    oas = []
                    aas = []
                    kappas = []
                    for index in config["index"]:  # 1-5, 间隔1
                        # for index in [1, 2, 3, 4, 5]:  # 1-5, 间隔1
                        Index = index
                        MARGIN = int((Patch_Size - 1) / 2)
                        num_tokens = (Patch_Size - 4) ** 2
                        print("Data set:", DATA_SET_NAME, "   ", end="")
                        print("PATCH_SIZE:", PATCH_SIZE, "   ", end="")
                        print("BATCH_SIZE:", BATCH_SIZE_TRAIN, "   ", end="")
                        print("learning_rate:", learning_rate, "   ", end="")
                        print("test_ratio", test_ratio, "   ", end="")
                        print("index:", "   ", Index, )

                        oa, aa, kappa = main()

                        oas.append(oa)
                        aas.append(aa)
                        kappas.append(kappa)

                    # 计算平均值
                    oa_avg = np.mean(oas)
                    aa_avg = np.mean(aas)
                    kappa_avg = np.mean(kappas)

                    parameters = {
                        "Module_Name": Module_Name,
                        "DATA_SET_NAME": DATA_SET_NAME,
                        "BATCH_SIZE_TRAIN": BATCH_SIZE_TRAIN,
                        "PATCH_SIZE": PATCH_SIZE,
                        "learning_rate": learning_rate,
                        # "T_max": T_max,
                        # "pca_components": pca_components,
                        "test_ratio": test_ratio,
                    }
                    # save_result_to_excel(oas, aas, kappas, parameters=parameters)
                    save_result_to_excel(oas, aas, kappas, parameters, Folder_Name, Module_Name, DATA_SET_NAME)
                    # 打印结果
                    print(f">>>>>> >>>>>> >>>>>>>>> total average OA: {oa_avg:.4f}", f"AA: {aa_avg:.4f}",
                          f"Kappa: {kappa_avg:.4f}\n\n")

